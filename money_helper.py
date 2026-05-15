#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import glob
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable, Iterable


DATE_HEADERS = {"date", "transaction date", "trans. date"}
AMOUNT_HEADER = "amount"
DEFAULT_CATEGORY_TEXT = "# Input\n- Others\n# Output\n- Others\n"
TOKEN_RE = re.compile(r"'[^']*'|\bOR\b|&|[^,&]+", re.IGNORECASE)
PRINT_MODE = 1


def log_message(message: str, level: int = 0) -> None:
    if PRINT_MODE == 1 and PRINT_MODE >= level:
        print(f"[money_helper] {message}")


def log_summary(message: str) -> None:
    if PRINT_MODE == 0:
        print(f"[money_helper] {message}")


@dataclass
class Rule:
    display_name: str
    raw_text: str
    matcher: Callable[[str], bool]
    is_others: bool = False


@dataclass
class FileData:
    path: Path
    file_type: str
    raw_row_count: int
    header_row_index: int
    header: list[str]
    head_rows: list[list[str]]
    rows: list[dict[str, str]]


@dataclass
class PrintedRow:
    section_name: str
    category_name: str
    rule_raw_text: str
    is_others: bool
    row: dict[str, str]


def error(message: str) -> None:
    raise RuntimeError(message)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description="""goal: Summarize bank statement CSV or Excel files by income and spending category.

step-by-step algorithm:
  Step 1. Read input file paths, date filters, category rules, row-print filter, and print mode.
  Step 2. Expand input paths or wildcards and load the markdown category rules.
  Step 3. For each input file, detect the header row, parse dates and amounts, and skip unusable rows.
  Step 4. Use each amount sign to choose Input or Output rules, then assign the first matching category.
  Step 5. Print per-file descriptions, selected rows, category totals, validation checks, and a cross-file final summary.

outputs:
  The script uses --input, --start_date, --end_date, --category, and --print to print summaries to stdout.
  No output file is written. Each printed artifact is generated from the matched rows in --input after the required date and category filters are applied.
""",
    )
    parser._optionals.title = "options"
    required = parser.add_argument_group("required input arguments")
    optional = parser.add_argument_group("optional input arguments")
    required.add_argument(
        "-i",
        "--input",
        nargs="+",
        required=True,
        metavar="INPUT_FILE",
        help="Required. Input CSV or Excel file path(s), provided with -i or --input; shell-style wildcards are supported.",
    )
    required.add_argument(
        "--start_date",
        required=True,
        metavar="YYYY-MM-DD",
        help="Required. Inclusive start date used to filter transaction rows, format YYYY-MM-DD.",
    )
    required.add_argument(
        "--end_date",
        required=True,
        metavar="YYYY-MM-DD",
        help="Required. Inclusive end date used to filter transaction rows, format YYYY-MM-DD.",
    )
    required.add_argument(
        "--category",
        required=True,
        metavar="CATEGORY_MD",
        help="Required. Markdown category definition file with # Input and # Output sections.",
    )
    optional.add_argument(
        "--print",
        dest="print_filter",
        default="All",
        metavar="ROW_FILTER",
        help="Optional. Row-print filter. Default: All. Special values: Others and All. Other values match comma-separated keywords or quoted phrases in row text.",
    )
    optional.add_argument(
        "--print_mode",
        dest="print_mode",
        type=int,
        choices=(0, 1),
        default=1,
        metavar="PRINT_MODE",
        help="Optional. Printing mode: 0 prints the final summary only, 1 prints all step details and report sections. Default: 1.",
    )
    return parser.parse_args()


def parse_iso_date(value: str, name: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        error(f"Invalid {name}: {value!r}. Expected YYYY-MM-DD.")
        raise exc


def parse_any_date(value: str) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in (
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y-%m-%d",
        "%m-%d-%Y",
        "%m/%d/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_amount(value: str) -> Decimal | None:
    text = str(value or "").strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    cleaned = text.replace("$", "").replace(",", "").strip("()").strip()
    try:
        amount = Decimal(cleaned)
    except InvalidOperation:
        return None
    return -amount if negative else amount


def is_supported_file(path: Path) -> bool:
    return path.suffix.lower() in {".csv", ".xlsx", ".xlsm", ".xls"}


def expand_inputs(patterns: Iterable[str]) -> list[Path]:
    matches: list[Path] = []
    for pattern in patterns:
        expanded = sorted(Path(p).resolve() for p in glob.glob(pattern))
        if expanded:
            matches.extend(expanded)
            continue
        path = Path(pattern).resolve()
        if path.exists():
            matches.append(path)
    deduped: list[Path] = []
    seen: set[Path] = set()
    for path in matches:
        if path not in seen:
            seen.add(path)
            deduped.append(path)
    if not deduped:
        error("No input files matched the provided pattern(s).")
    unsupported = [str(path) for path in deduped if not is_supported_file(path)]
    if unsupported:
        error(f"Only CSV or Excel files are supported. Unsupported file(s): {', '.join(unsupported)}")
    return deduped


def load_tabular_rows(path: Path) -> list[list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            return [[str(cell).strip() for cell in row] for row in csv.reader(handle)]
    if suffix == ".xls":
        error(f"Legacy .xls is not supported for {path}. Please convert it to .xlsx or .csv.")
    try:
        from openpyxl import load_workbook
    except ImportError:
        error(f"Reading Excel files requires openpyxl, but it is not installed: {path}")
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell).strip() for cell in row])
    return rows


def normalize_header(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def row_is_empty(row: list[str]) -> bool:
    return all(not str(cell).strip() for cell in row)


def first_nonempty_cell(row: list[str]) -> str:
    for cell in row:
        text = str(cell).strip()
        if text:
            return text
    return ""


def detect_header(rows: list[list[str]], path: Path) -> tuple[int, list[str], list[list[str]]]:
    for idx, row in enumerate(rows):
        normalized = [normalize_header(cell) for cell in row]
        if AMOUNT_HEADER not in normalized:
            continue
        if not any(name in normalized for name in DATE_HEADERS):
            continue
        header = [str(cell).strip() for cell in row]
        head_rows = [r for r in rows[:idx] if not row_is_empty(r)]
        return idx, header, head_rows
    error(f"Could not find a header row with both date and amount columns in {path}")
    raise AssertionError


def load_file_data(path: Path) -> FileData:
    raw_rows = load_tabular_rows(path)
    header_row_idx, header, head_rows = detect_header(raw_rows, path)
    normalized = [normalize_header(name) for name in header]
    amount_idx = normalized.index(AMOUNT_HEADER)
    date_indexes = [i for i, name in enumerate(normalized) if name in DATE_HEADERS]
    if amount_idx < 0:
        error(f"Missing Amount column in {path}")
    if not date_indexes:
        error(f"Missing date column in {path}")

    data_rows: list[dict[str, str]] = []
    for row in raw_rows[header_row_idx + 1 :]:
        if row_is_empty(row):
            continue
        padded = row + [""] * (len(header) - len(row))
        data_rows.append({header[i]: padded[i].strip() for i in range(len(header))})
    return FileData(
        path=path,
        file_type=path.suffix.lower().lstrip(".").upper(),
        raw_row_count=len(data_rows),
        header_row_index=header_row_idx + 1,
        header=header,
        head_rows=head_rows,
        rows=data_rows,
    )


def split_unquoted_commas(text: str) -> list[str]:
    parts: list[str] = []
    current: list[str] = []
    in_quote = False
    for ch in text:
        if ch == "'":
            in_quote = not in_quote
        if ch == "," and not in_quote:
            part = "".join(current).strip()
            if part:
                parts.append(part)
            current = []
        else:
            current.append(ch)
    part = "".join(current).strip()
    if part:
        parts.append(part)
    return parts


def parse_term(token: str, raw_expr: str) -> str:
    term = token.strip()
    if not term:
        error(f"Invalid empty term in category rule: {raw_expr!r}")
    if term.startswith("'") and term.endswith("'"):
        inner = term[1:-1]
        if not inner.strip():
            error(f"Invalid empty quoted term in category rule: {raw_expr!r}")
        return inner.lower()
    if " " in term:
        error(f"Unquoted term contains spaces in category rule: {raw_expr!r}")
    return term.lower()


def clean_display_name(text: str) -> str:
    return re.sub(r"'([^']*)'", r"\1", text).strip()


def build_matcher(expr: str):
    tokens = [tok.strip() for tok in TOKEN_RE.findall(expr) if tok.strip()]
    if not tokens:
        error(f"Invalid empty category rule: {expr!r}")
    terms: list[str] = []
    operators: list[str] = []
    expecting_term = True
    for token in tokens:
        upper = token.upper()
        if upper == "OR" or token == "&":
            if expecting_term:
                error(f"Category rule cannot start with an operator: {expr!r}")
            operators.append(upper)
            expecting_term = True
            continue
        if not expecting_term:
            error(f"Missing operator in category rule: {expr!r}")
        terms.append(parse_term(token, expr))
        expecting_term = False
    if expecting_term:
        error(f"Category rule cannot end with an operator: {expr!r}")
    if not operators:
        needle = terms[0]
        return lambda text: needle in text
    if len(set(operators)) != 1:
        error(f"Do not mix '&' and 'OR' in the same category rule: {expr!r}")
    if operators[0] == "&":
        return lambda text: all(term in text for term in terms)
    return lambda text: any(term in text for term in terms)


def parse_category_file(path: str | None) -> dict[str, list[Rule]]:
    text = Path(path).read_text(encoding="utf-8") if path else DEFAULT_CATEGORY_TEXT
    sections = {"input": [], "output": []}
    current: str | None = None
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        lowered = stripped.lower()
        if lowered == "# input":
            current = "input"
            continue
        if lowered == "# output":
            current = "output"
            continue
        if not stripped.startswith("-"):
            continue
        if current is None:
            error("Category file must define # Input or # Output before list items.")
        raw_item = stripped[1:].strip().rstrip(",").strip()
        if not raw_item:
            continue
        if raw_item.lower() == "others":
            sections[current].append(Rule(display_name="Others", raw_text="Others", matcher=lambda _: True, is_others=True))
            continue
        expressions = split_unquoted_commas(raw_item)
        matchers = [build_matcher(expr) for expr in expressions]
        sections[current].append(
            Rule(
                display_name=clean_display_name(", ".join(expr.strip() for expr in expressions)),
                raw_text=raw_item,
                matcher=lambda text, ms=matchers: any(m(text) for m in ms),
                is_others=False,
            )
        )
    for section_name in ("input", "output"):
        others = [rule for rule in sections[section_name] if rule.is_others]
        normal = [rule for rule in sections[section_name] if not rule.is_others]
        if len(others) > 1:
            error(f"Section {section_name!r} has more than one Others category.")
        sections[section_name] = normal + others
        if not sections[section_name]:
            sections[section_name] = [Rule(display_name="Others", raw_text="Others", matcher=lambda _: True, is_others=True)]
    return sections


def resolve_row_date(row: dict[str, str], header: list[str]) -> date | None:
    lookup = {normalize_header(name): name for name in header}
    for header_name in header:
        normalized = normalize_header(header_name)
        if normalized not in DATE_HEADERS:
            continue
        return parse_any_date(row.get(lookup[normalized], ""))
    return None


def resolve_row_amount(row: dict[str, str], header: list[str]) -> Decimal | None:
    lookup = {normalize_header(name): name for name in header}
    header_name = lookup.get("amount")
    if header_name is None:
        return None
    raw_value = row.get(header_name, "")
    if not str(raw_value).strip():
        return None
    return parse_amount(raw_value)


def normalize_text(text: str) -> str:
    return " ".join(text.lower().split())


def parse_print_terms(print_filter: str | None) -> list[str]:
    if not print_filter:
        return []
    terms: list[str] = []
    for part in split_unquoted_commas(print_filter):
        cleaned = part.strip().strip("'").strip('"')
        if cleaned:
            terms.append(normalize_text(cleaned))
    return terms


def assign_category(row_text: str, rules: list[Rule], section_name: str, path: Path, row_number: int) -> Rule:
    matches = [rule for rule in rules if not rule.is_others and rule.matcher(row_text)]
    if matches:
        return matches[0]
    for rule in rules:
        if rule.is_others:
            return rule
    error(f"No {section_name} category matched row {row_number} in {path}, and Others is not defined.")
    raise AssertionError


def money(value: Decimal) -> str:
    return f"{value:,.2f}"


def pct(value: Decimal, total: Decimal) -> str:
    if total == 0:
        return "0.00%"
    return f"{(value / total * Decimal('100')):.2f}%"


def summarize_file(
    file_data: FileData,
    category_rules: dict[str, list[Rule]],
    start_date: date | None,
    end_date: date | None,
    print_filter: str | None,
):
    summary: dict[str, dict[str, dict[str, Decimal | int]]] = {"input": {}, "output": {}}
    matched_rows = 0
    balance = Decimal("0")
    seen_dates: list[date] = []
    printed_rows: list[PrintedRow] = []
    print_terms = parse_print_terms(print_filter)
    print_all_rows = len(print_terms) == 1 and print_terms[0] == "all"
    print_others_only = len(print_terms) == 1 and print_terms[0] == "others"
    skipped_invalid_amount = 0
    skipped_invalid_date = 0
    skipped_before_start = 0
    skipped_after_end = 0
    skipped_zero_amount = 0
    example_named: PrintedRow | None = None
    example_others: PrintedRow | None = None
    amount_example: tuple[str, Decimal, str] | None = None
    date_column_used = next((name for name in file_data.header if normalize_header(name) in DATE_HEADERS), None)

    for section_name, rules in category_rules.items():
        for rule in rules:
            summary[section_name][rule.display_name] = {"count": 0, "sum": Decimal("0")}

    for idx, row in enumerate(file_data.rows, start=1):
        amount = resolve_row_amount(row, file_data.header)
        if amount is None:
            skipped_invalid_amount += 1
            continue
        row_date = resolve_row_date(row, file_data.header)
        if row_date is None:
            skipped_invalid_date += 1
            continue
        seen_dates.append(row_date)
        if start_date and row_date < start_date:
            skipped_before_start += 1
            continue
        if end_date and row_date > end_date:
            skipped_after_end += 1
            continue
        if amount == 0:
            skipped_zero_amount += 1
            continue
        section_name = "input" if amount > 0 else "output"
        row_text = normalize_text(" ".join(str(value) for value in row.values() if str(value).strip()))
        rule = assign_category(row_text, category_rules[section_name], section_name, file_data.path, idx)
        bucket = summary[section_name][rule.display_name]
        bucket["count"] = int(bucket["count"]) + 1
        bucket["sum"] = Decimal(bucket["sum"]) + abs(amount)
        if amount_example is None:
            amount_example = (str(row.get("Amount", "")), abs(amount), section_name.title())
        if print_all_rows:
            printed_rows.append(
                PrintedRow(
                    section_name=section_name,
                    category_name=rule.display_name,
                    rule_raw_text=rule.raw_text,
                    is_others=rule.is_others,
                    row=row,
                )
            )
        elif print_others_only and rule.is_others:
            printed_rows.append(
                PrintedRow(
                    section_name=section_name,
                    category_name=rule.display_name,
                    rule_raw_text=rule.raw_text,
                    is_others=rule.is_others,
                    row=row,
                )
            )
        elif print_terms and any(term in row_text for term in print_terms):
            printed_rows.append(
                PrintedRow(
                    section_name=section_name,
                    category_name=rule.display_name,
                    rule_raw_text=rule.raw_text,
                    is_others=rule.is_others,
                    row=row,
                )
            )
        example_row = PrintedRow(
            section_name=section_name,
            category_name=rule.display_name,
            rule_raw_text=rule.raw_text,
            is_others=rule.is_others,
            row=row,
        )
        if rule.is_others and example_others is None:
            example_others = example_row
        if not rule.is_others and example_named is None:
            example_named = example_row
        matched_rows += 1
        balance += amount

    effective_start = start_date or (min(seen_dates) if seen_dates else None)
    effective_end = end_date or (max(seen_dates) if seen_dates else None)
    return {
        "summary": summary,
        "matched_rows": matched_rows,
        "balance": balance,
        "effective_start": effective_start,
        "effective_end": effective_end,
        "printed_rows": printed_rows,
        "date_column_used": date_column_used,
        "skipped_invalid_amount": skipped_invalid_amount,
        "skipped_invalid_date": skipped_invalid_date,
        "skipped_before_start": skipped_before_start,
        "skipped_after_end": skipped_after_end,
        "skipped_zero_amount": skipped_zero_amount,
        "example_named": example_named,
        "example_others": example_others,
        "amount_example": amount_example,
    }


def ordinal(n: int) -> str:
    if 10 <= n % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def format_statement_date(value: date | None) -> str:
    return value.isoformat() if value else "Unknown"


def format_summary(
    title: str,
    matched_rows: int,
    total_balance: Decimal,
    summary: dict[str, dict[str, dict[str, Decimal | int]]],
    processed_files: list[Path] | None = None,
) -> str:
    input_named = {k: v for k, v in summary["input"].items() if k != "Others"}
    output_named = {k: v for k, v in summary["output"].items() if k != "Others"}
    input_others = summary["input"].get("Others", {"count": 0, "sum": Decimal("0")})
    output_others = summary["output"].get("Others", {"count": 0, "sum": Decimal("0")})

    total_input_named = sum(Decimal(v["sum"]) for v in input_named.values())
    total_output_named = sum(Decimal(v["sum"]) for v in output_named.values())
    input_others_sum = Decimal(input_others["sum"])
    output_others_sum = Decimal(output_others["sum"])
    others_net = input_others_sum - output_others_sum
    row_parts: list[str] = []
    balance_parts: list[str] = []
    row_total = 0
    balance_total = Decimal("0")

    lines = [
        title,
    ]

    if processed_files is None:
        lines.append("Processed data file: 1")
    else:
        lines.append("Processed data files:")
        for index, path in enumerate(processed_files, start=1):
            lines.append(f"{index}. {path}")

    lines.extend(
        [
        f"",
        f"Matched Rows: {matched_rows}",
        f"Total Balance ($) = Input - Output + Others = {money(total_input_named)} - {money(total_output_named)} + {money(others_net)} = {money(total_balance)}",
        "",
        f"Input by Category ($): {money(total_input_named)}",
        ]
    )

    for name, values in input_named.items():
        count = int(values["count"])
        amount = Decimal(values["sum"])
        lines.append(f"  {name} ({count}): {money(amount)} ({pct(amount, total_input_named)})")
        row_parts.append(str(count))
        balance_parts.append(money(amount))
        row_total += count
        balance_total += amount

    if not input_named:
        lines.append("  None")

    lines.extend(["", f"Output by Category ($): {money(total_output_named)}"])
    for name, values in output_named.items():
        count = int(values["count"])
        amount = Decimal(values["sum"])
        lines.append(f"  {name} ({count}): {money(amount)} ({pct(amount, total_output_named)})")
        row_parts.append(str(count))
        balance_parts.append(money(-amount))
        row_total += count
        balance_total -= amount

    if not output_named:
        lines.append("  None")

    input_others_count = int(input_others["count"])
    output_others_count = int(output_others["count"])
    row_parts.extend([str(input_others_count), str(output_others_count)])
    balance_parts.extend([money(input_others_sum), money(-output_others_sum)])
    row_total += input_others_count + output_others_count
    balance_total += input_others_sum - output_others_sum

    lines.extend(
        [
            "",
            f"Others ($): {money(others_net)}",
            f"  Input Others ({input_others_count}): {money(input_others_sum)}",
            f"  Output Others ({output_others_count}): {money(output_others_sum)}",
            f"  Input - Output: {money(input_others_sum)} - {money(output_others_sum)} = {money(input_others_sum - output_others_sum)}",
            "",
            "Checks:",
            f"  Row count check: {' + '.join(row_parts)} = {row_total} --> {'PASS' if row_total == matched_rows else 'FAIL'}",
            f"  Balance check: {' + '.join(balance_parts)} = {money(balance_total)} --> {'PASS' if balance_total == total_balance else 'FAIL'}",
        ]
    )
    return "\n".join(lines)


def merge_summaries(target: dict[str, dict[str, dict[str, Decimal | int]]], source: dict[str, dict[str, dict[str, Decimal | int]]]) -> None:
    for section_name, categories in source.items():
        for category_name, values in categories.items():
            bucket = target[section_name].setdefault(category_name, {"count": 0, "sum": Decimal("0")})
            bucket["count"] = int(bucket["count"]) + int(values["count"])
            bucket["sum"] = Decimal(bucket["sum"]) + Decimal(values["sum"])


def format_head_information(head_rows: list[list[str]]) -> list[str]:
    if not head_rows:
        return ["Head Information: None"]
    lines = ["Head Information:"]
    for row in head_rows:
        lines.append(f"  {', '.join(cell for cell in row if cell)}")
    return lines


def truncate_text(text: str, limit: int = 80) -> str:
    compact = " ".join(str(text).split())
    return compact if len(compact) <= limit else compact[: limit - 3] + "..."


def matching_rule_details(raw_rule_text: str, is_others: bool, row_text: str) -> str:
    if is_others:
        return "No named category rule matched this normalized row text, so the fallback Others bucket was used."

    matched_expressions = [expr.strip() for expr in split_unquoted_commas(raw_rule_text) if build_matcher(expr)(row_text)]
    if matched_expressions:
        return f"Matched named rule {raw_rule_text!r} because the normalized row text satisfied: {', '.join(clean_display_name(expr) for expr in matched_expressions)}."
    return f"Assigned by named rule {raw_rule_text!r}."


def format_example_row(item: PrintedRow | None) -> str:
    if item is None:
        return "None"
    row = item.row
    return ", ".join(f"{key}={value}" for key, value in row.items() if str(value).strip())


def get_example_details(item: PrintedRow | None, date_column: str | None) -> list[str]:
    if item is None:
        return ["    None"]
    row = item.row
    normalized_date_column = normalize_header(date_column or "")
    date_value = next((value for key, value in row.items() if normalize_header(key) == normalized_date_column and value), "")
    amount_value = next((value for key, value in row.items() if normalize_header(key) == "amount" and value), "")
    description = next((value for key, value in row.items() if normalize_header(key) == "description" and value), "")
    amount = parse_amount(amount_value) if amount_value else None
    row_text = normalize_text(" ".join(str(value) for value in row.values() if str(value).strip()))
    amount_detail = "Amount could not be parsed."
    if amount is not None:
        direction = "Input" if amount > 0 else "Output"
        amount_detail = f"Amount handling: raw Amount={amount_value} -> {direction} amount {money(abs(amount))}."
    category_detail = (
        f"Category match: Amount first determined the section as {item.section_name.title()}, "
        f"then the row was assigned to {item.section_name.title()} -> {item.category_name}."
    )
    if description:
        category_detail += f" Matching used the normalized row text, including Description={truncate_text(description)}."
    category_reason = "Category reason: "
    category_reason += matching_rule_details(item.rule_raw_text, item.is_others, row_text)
    date_detail = f"Date handling: used {date_column or 'Unknown'}={date_value or 'Unknown'} as the true date."
    return [
        f"    Row: {format_example_row(item)}",
        f"    {date_detail}",
        f"    {amount_detail}",
        f"    {category_detail}",
        f"    {category_reason}",
    ]


def format_description(
    file_data: FileData,
    result: dict[str, object],
    start_date: date | None,
    end_date: date | None,
) -> list[str]:
    effective_start = format_statement_date(result["effective_start"])
    effective_end = format_statement_date(result["effective_end"])
    date_column = result["date_column_used"] or "Unknown"
    skipped_before_start = int(result["skipped_before_start"])
    skipped_after_end = int(result["skipped_after_end"])
    skipped_invalid_amount = int(result["skipped_invalid_amount"])
    skipped_invalid_date = int(result["skipped_invalid_date"])
    skipped_zero_amount = int(result["skipped_zero_amount"])
    matched_rows = int(result["matched_rows"])
    skipped_total = skipped_before_start + skipped_after_end + skipped_invalid_amount + skipped_invalid_date + skipped_zero_amount
    amount_example = result["amount_example"]
    amount_text = "No matched amount example."
    if amount_example is not None:
        raw_amount, absolute_amount, section_name = amount_example
        amount_text = f"Example amount conversion: raw Amount={raw_amount} -> {section_name} amount {money(absolute_amount)}."
    filter_text = (
        f"Included rows whose parsed {date_column} was between {start_date.isoformat()} and {end_date.isoformat()} (inclusive)."
        if start_date and end_date
        else f"Used all rows whose parsed {date_column} fell within the detected statement range {effective_start} to {effective_end}."
    )
    return [
        "Description:",
        f"  File Type: {file_data.file_type}",
        f"  Header Detection: Found the header at row {file_data.header_row_index} with date and amount columns; {len(file_data.head_rows)} head rows were treated as Head Information.",
        f"  Date Handling: Used {date_column} as the true date; statement dates in this file range from {effective_start} to {effective_end}.",
        "  Amount Handling: Positive amounts were counted as Input; negative amounts were counted as Output; zero-amount rows were ignored. The amount sign decides whether the row is matched against Input rules or Output rules.",
        f"  Row Processing: {filter_text} {file_data.raw_row_count} raw rows -> {matched_rows} matched rows; {skipped_total} rows were skipped ({skipped_before_start} before start date, {skipped_after_end} after end date, {skipped_invalid_amount} invalid/missing amount, {skipped_invalid_date} invalid/missing date, {skipped_zero_amount} zero amount).",
        "  Category Matching: Each row first used its Amount sign to choose the Input or Output section, then the row was matched against that section's markdown rules in order; the first named match was used, and unmatched rows were assigned to Others.",
        "  Example named match:",
        *get_example_details(result["example_named"], date_column),
        "  Example Others match:",
        *get_example_details(result["example_others"], date_column),
        f"  Additional amount example: {amount_text}",
        "  Summary Contribution: This file's category totals were added to the cross-file final summary.",
    ]


def format_printed_rows(label: str, printed_rows: list[PrintedRow]) -> list[str]:
    if not printed_rows:
        return [f"Printed Rows ({label}): None"]
    lines = [f"Printed Rows ({label}):"]
    for item in printed_rows:
        row_text = ", ".join(f"{key}={value}" for key, value in item.row.items() if str(value).strip())
        lines.append(f"  [{item.section_name.title()} {item.category_name}] {row_text}")
    return lines


def print_detailed_report(
    input_paths: list[Path],
    category_rules: dict[str, list[Rule]],
    start_date: date | None,
    end_date: date | None,
    print_filter: str | None,
) -> tuple[int, Decimal]:
    log_message("Step 3: processing input files", level=1)
    overall_summary = {
        "input": {rule.display_name: {"count": 0, "sum": Decimal("0")} for rule in category_rules["input"]},
        "output": {rule.display_name: {"count": 0, "sum": Decimal("0")} for rule in category_rules["output"]},
    }
    overall_rows = 0
    overall_balance = Decimal("0")

    for file_index, path in enumerate(input_paths, start=1):
        log_message(f"Processing file {file_index} of {len(input_paths)}: {path}", level=1)
        file_data = load_file_data(path)
        result = summarize_file(file_data, category_rules, start_date, end_date, print_filter)
        merge_summaries(overall_summary, result["summary"])
        overall_rows += result["matched_rows"]
        overall_balance += result["balance"]

        if PRINT_MODE == 1:
            print(f"# Processing the {ordinal(file_index)} data file: {path}")
            print("```text")
            for line in format_description(file_data, result, start_date, end_date):
                print(line)
            for line in format_head_information(file_data.head_rows):
                print(line)
            if print_filter or result["printed_rows"]:
                print_label = print_filter or "Matched"
                for line in format_printed_rows(print_label, result["printed_rows"]):
                    print(line)
            print()
            print(format_summary("Data File Summary", result["matched_rows"], result["balance"], result["summary"]))
            print("```")
            print()

    log_message("Step 4: printing final summary", level=1)
    print("# Final Summary")
    print("```text")
    print("Description: This final summary is the sum of all per-file category totals after the same date, amount, and category-matching rules were applied to each file.")
    print(format_summary("", overall_rows, overall_balance, overall_summary, processed_files=input_paths))
    print("```")

    return overall_rows, overall_balance


def main() -> int:
    args = parse_args()
    global PRINT_MODE
    PRINT_MODE = args.print_mode

    try:
        log_message("Step 1: reading input arguments", level=1)
        log_message(
            "Inputs understood as: "
            f"input={args.input!r}, start_date={args.start_date!r}, end_date={args.end_date!r}, "
            f"category={args.category!r}, print={args.print_filter!r}, print_mode={args.print_mode}",
            level=0,
        )

        start_date = parse_iso_date(args.start_date, "start_date") if args.start_date else None
        end_date = parse_iso_date(args.end_date, "end_date") if args.end_date else None
        if start_date and end_date and start_date > end_date:
            error("start_date cannot be later than end_date.")

        log_message("Step 2: loading input files and category rules", level=1)
        input_paths = expand_inputs(args.input)
        category_rules = parse_category_file(args.category)
        log_message(
            f"Loaded {len(input_paths)} input file(s), {len(category_rules['input'])} input rule(s), "
            f"and {len(category_rules['output'])} output rule(s).",
            level=0,
        )

        overall_rows, overall_balance = print_detailed_report(
            input_paths=input_paths,
            category_rules=category_rules,
            start_date=start_date,
            end_date=end_date,
            print_filter=args.print_filter,
        )
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    log_message("Step 5. Finished Successfully", level=1)
    log_message(
        f"Final summary: processed_files={len(input_paths)}, matched_rows={overall_rows}, total_balance={money(overall_balance)}",
        level=0,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
